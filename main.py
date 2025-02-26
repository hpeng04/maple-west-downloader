import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from unit import Unit
from color import color
from channels import channels
import calendar
import pandas as pd
import re
import zipfile
from pathlib import Path
import time
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class QualityChecker:
    def __init__(self, config_path='config/', output_dir=None):
        self.config_path = config_path
        self.output_dir = output_dir or tempfile.mkdtemp(prefix="maple_west_quality_")
        self.units = self._load_units()
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
    def _load_units(self):
        """Load all unit configurations from the config directory"""
        units = {}
        for file in os.listdir(self.config_path):
            if file.endswith('.json'):
                with open(os.path.join(self.config_path, file), 'r') as f:
                    config = json.load(f)
                    units[config['unit_no']] = config
        return units

    def _check_data_quality(self, data, unit_config, channel_name):
        """Check data quality for a specific channel"""
        if not unit_config['channels'].get(channel_name, False):
            return None  # Channel not monitored for this unit
            
        channel_info = channels[channel_name]
        values = data[data.columns[data.columns.str.contains(channel_info.regex, regex=True)]]
        
        if values.empty:
            return None  # Channel not found in data
            
        values = values.iloc[:, 0]  # Take first matching column
        
        good_count = 0
        missing_count = 0
        bad_count = 0
        
        for value in values:
            if pd.isna(value) or value == "":
                missing_count += 1
            elif not isinstance(value, (int, float)):
                try:
                    value = float(value)
                except:
                    missing_count += 1
                    continue
                    
            if isinstance(value, (int, float)):
                if channel_info.min_value <= value <= channel_info.max_value:
                    good_count += 1
                else:
                    bad_count += 1
                    
        return [good_count, missing_count, bad_count]

    def _format_quality_result(self, counts):
        """Format quality results as text"""
        if counts is None or len(counts) != 3:
            return "Good: 0, Missing: 0, Bad: 0"
        return f"Good: {counts[0]}, Missing: {counts[1]}, Bad: {counts[2]}"

    def _apply_conditional_formatting(self, worksheet):
        """Apply conditional formatting to cells based on missing + bad percentage"""
        for row in worksheet.iter_rows(min_row=2):  # Skip header row
            for cell in row[1:]:  # Skip date column
                if cell.value:
                    # Extract counts from the formatted string
                    parts = cell.value.split(', ')
                    good = int(parts[0].split(': ')[1])
                    missing = int(parts[1].split(': ')[1])
                    bad = int(parts[2].split(': ')[1])
                    
                    total = good + missing + bad
                    if total > 0:
                        problem_percentage = (missing + bad) / total * 100
                        if problem_percentage > 5:
                            cell.fill = self.red_fill
                        elif problem_percentage > 1:
                            cell.fill = self.yellow_fill

    def process_data(self, data_path, unit_no, month, data_type):
        """Process data for a specific unit and month"""
        unit_config = self.units[unit_no]
        results = {}
        
        # Initialize results for all channels
        for channel_name in channels.keys():
            if unit_config['channels'].get(channel_name, False):
                results[channel_name] = [0, 0, 0]  # [good, missing, bad]
                
        # Read all CSV files for the unit and month
        unit_path = os.path.join(data_path, f'UNIT {unit_no}')
        if not os.path.exists(unit_path):
            print(f"No data directory found for Unit {unit_no} in {unit_path}")
            return None
            
        for file in os.listdir(unit_path):
            # Check if file matches the month and is a CSV file
            if month in file and file.endswith('.csv'):
                try:
                    data = pd.read_csv(os.path.join(unit_path, file))
                    
                    # Check quality for each channel
                    for channel_name in channels.keys():
                        if unit_config['channels'].get(channel_name, False):
                            quality = self._check_data_quality(data, unit_config, channel_name)
                            if quality:
                                results[channel_name] = [x + y for x, y in zip(results[channel_name], quality)]
                except Exception as e:
                    print(f"Error processing file {file}: {str(e)}")
                    continue
                            
        return results

    def update_quality_report(self, data_type='Minute'):
        """Update quality reports for all units"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir, exist_ok=True)
            
        for unit_no in self.units.keys():
            unit_path = os.path.join(self.output_dir, f'UNIT {unit_no}')
            if not os.path.exists(unit_path):
                print(f"No data directory found for Unit {unit_no}")
                continue

            report_file = os.path.join(self.output_dir, 'quality_reports', f'unit_{unit_no}_{data_type.lower()}_quality.xlsx')
            os.makedirs(os.path.dirname(report_file), exist_ok=True)
            
            # Load existing report or create new one
            if os.path.exists(report_file):
                report_df = pd.read_excel(report_file, index_col=0)
            else:
                report_df = pd.DataFrame(columns=[ch for ch in channels.keys() if self.units[unit_no]['channels'].get(ch, False)])
            
            # Get unique months from filenames
            months = set()
            for file in os.listdir(unit_path):
                if data_type in file and file.endswith('.csv'):
                    try:
                        # Extract YYYY-MM from filename
                        date_str = file.split('_')[-1].replace('.csv', '')
                        if len(date_str) >= 7:  # Ensure we have at least YYYY-MM
                            months.add(date_str[:7])  # Get YYYY-MM part
                    except:
                        continue
            
            # Process each month's data
            new_data_added = False
            for month in months:
                if month not in report_df.index:
                    results = self.process_data(self.output_dir, unit_no, month, data_type)
                    if results:
                        report_df.loc[month] = {ch: self._format_quality_result(results.get(ch)) for ch in report_df.columns}
                        new_data_added = True
                else:
                    print(f"Month {month} already exists in report for Unit {unit_no}")
            
            if not new_data_added:
                print(f"No new data to add for Unit {unit_no}")
                continue
            
            # Save updated report with conditional formatting
            report_df.sort_index().to_excel(report_file)
            
            # Apply conditional formatting
            wb = load_workbook(report_file)
            ws = wb.active
            self._apply_conditional_formatting(ws)
            wb.save(report_file)
            
            print(f"Updated quality report for Unit {unit_no} ({data_type} data)")

class BulkDownloadGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Maple West Bulk Download")
        self.root.geometry("800x600")  # Set initial window size
        
        # Configure grid weights for main window
        self.root.grid_columnconfigure(0, weight=3)  # Left side (unit selection, etc)
        self.root.grid_columnconfigure(1, weight=2)  # Right side (download controls)
        self.root.grid_rowconfigure(0, weight=1)
        
        # Create left frame for selections
        self.left_frame = ttk.Frame(self.root, padding="10")
        self.left_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configure left frame grid
        self.left_frame.grid_columnconfigure(0, weight=1)
        for i in range(5):  # For unit selection, date selection, data type, combine option, quality option
            self.left_frame.grid_rowconfigure(i, weight=1)
        
        # Create right frame for download controls
        self.right_frame = ttk.Frame(self.root, padding="10")
        self.right_frame.grid(row=0, column=1, sticky="nsew")
        
        # Configure right frame grid
        self.right_frame.grid_columnconfigure(0, weight=1)
        self.right_frame.grid_rowconfigure(0, weight=1)  # Space above
        self.right_frame.grid_rowconfigure(1, weight=0)  # Download controls
        self.right_frame.grid_rowconfigure(2, weight=1)  # Space below
        
        # Initialize variables
        self.cancel_flag = False
        self.temp_dirs = []
        self.units = self.load_units('config/')  # Load units first
        self.unit_vars = {}
        self.select_all_var = tk.BooleanVar(value=False)  # Add select all variable
        self.combine_var = tk.BooleanVar(value=True)
        self.quality_report_var = tk.BooleanVar(value=True)
        self.is_downloading = False
        self.start_time = None  # Add start_time initialization
        
        # Create GUI elements
        self.create_unit_selection()
        self.create_date_selection()
        self.create_data_type_selection()
        self.create_combine_option()
        self.create_quality_option()
        self.create_download_controls()
        
        # Set minimum window size
        self.root.minsize(800, 600)

    def create_unit_selection(self):
        # Unit selection frame
        unit_frame = ttk.LabelFrame(self.left_frame, text="Unit Selection", padding="5")
        unit_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # Configure unit frame grid
        unit_frame.grid_columnconfigure(0, weight=1)
        unit_frame.grid_columnconfigure(1, weight=1)
        
        # Select All checkbox
        select_all_cb = ttk.Checkbutton(unit_frame, text="Select All", 
                                       variable=self.select_all_var,
                                       command=self.toggle_all_units)
        select_all_cb.grid(row=0, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        
        # Create scrollable frame for units
        canvas = tk.Canvas(unit_frame, height=200)  # Set a fixed height
        scrollbar = ttk.Scrollbar(unit_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Bind mouse wheel to scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create unit checkboxes
        row = 0
        col = 0
        max_rows = (len(self.units) + 1) // 2  # Ensure even distribution in columns
        for unit in sorted(self.units, key=lambda x: x.unit_no):
            var = tk.BooleanVar()
            self.unit_vars[unit.unit_no] = var
            cb = ttk.Checkbutton(scrollable_frame, text=f"Unit {unit.unit_no}", 
                                variable=var)
            cb.grid(row=row, column=col, sticky="w", padx=5, pady=2)
            row += 1
            if row >= max_rows:
                row = 0
                col = 1
        
        # Grid the canvas and scrollbar
        canvas.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5)
        scrollbar.grid(row=1, column=2, sticky="ns")
        unit_frame.grid_rowconfigure(1, weight=1)

    def create_date_selection(self):
        # Date selection frame
        date_frame = ttk.LabelFrame(self.left_frame, text="Select Date Range", padding="5")
        date_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=5)
        
        # Start date
        ttk.Label(date_frame, text="Start:").grid(row=0, column=0, padx=5)
        
        # Create month and year selection for start date
        self.start_month = ttk.Combobox(date_frame, values=list(calendar.month_name)[1:], width=10)
        self.start_month.set(calendar.month_name[(datetime.now() - relativedelta(months=1)).month])
        self.start_month.grid(row=0, column=1, padx=5)
        
        current_year = datetime.now().year
        self.start_year = ttk.Combobox(date_frame, values=list(range(current_year-5, current_year+1)), width=6)
        self.start_year.set(str(current_year))
        self.start_year.grid(row=0, column=2, padx=5)
        
        # End date
        ttk.Label(date_frame, text="End:").grid(row=1, column=0, padx=5, pady=5)
        
        self.end_month = ttk.Combobox(date_frame, values=list(calendar.month_name)[1:], width=10)
        self.end_month.set(calendar.month_name[datetime.now().month])
        self.end_month.grid(row=1, column=1, padx=5)
        
        self.end_year = ttk.Combobox(date_frame, values=list(range(current_year-5, current_year+1)), width=6)
        self.end_year.set(str(current_year))
        self.end_year.grid(row=1, column=2, padx=5)

    def create_data_type_selection(self):
        # Data type selection frame
        type_frame = ttk.LabelFrame(self.left_frame, text="Select Data Type", padding="5")
        type_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=5)
        
        self.data_type = tk.StringVar(value="minute")
        ttk.Radiobutton(type_frame, text="Minute Data", variable=self.data_type,
                       value="minute").grid(row=0, column=0, padx=20)
        ttk.Radiobutton(type_frame, text="Hour Data", variable=self.data_type,
                       value="hour").grid(row=0, column=1, padx=20)

    def create_combine_option(self):
        # Combine data option frame
        combine_frame = ttk.LabelFrame(self.left_frame, text="Data Processing", padding="5")
        combine_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=5)
        
        ttk.Checkbutton(combine_frame, text="Combine downloaded data into single files", 
                       variable=self.combine_var).grid(row=0, column=0, padx=20)

    def create_quality_option(self):
        """Create quality report option"""
        quality_frame = ttk.LabelFrame(self.left_frame, text="Quality Report", padding="5")
        quality_frame.grid(row=4, column=0, columnspan=2, sticky="nsew", pady=5)
        
        ttk.Checkbutton(quality_frame, text="Generate quality report after download", 
                       variable=self.quality_report_var).grid(row=0, column=0, padx=20)

    def create_download_controls(self):
        # Create frame for download controls in the center of right frame
        download_frame = ttk.Frame(self.right_frame)
        download_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        # Configure download frame grid
        download_frame.grid_columnconfigure(0, weight=1)
        for i in range(4):
            download_frame.grid_rowconfigure(i, weight=0)
        
        # Download button
        self.download_btn = ttk.Button(download_frame, text="Start Download", command=self.start_download)
        self.download_btn.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        
        # Cancel button (hidden initially)
        self.cancel_btn = ttk.Button(download_frame, text="Cancel", command=self.cancel_download)
        self.cancel_btn.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        self.cancel_btn.grid_remove()
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(download_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
        
        # Progress frame (percentage and time remaining)
        progress_info_frame = ttk.Frame(download_frame)
        progress_info_frame.grid(row=3, column=0, sticky="ew", padx=5)
        progress_info_frame.grid_columnconfigure(0, weight=1)
        progress_info_frame.grid_columnconfigure(1, weight=1)
        
        self.percent_label = ttk.Label(progress_info_frame, text="0%")
        self.percent_label.grid(row=0, column=0, sticky="w")
        
        self.time_label = ttk.Label(progress_info_frame, text="")
        self.time_label.grid(row=0, column=1, sticky="e")
        
        # Status message
        self.status_label = ttk.Label(download_frame, text="", wraplength=250)
        self.status_label.grid(row=4, column=0, sticky="ew", padx=5, pady=5)

    def toggle_all_units(self):
        """Toggle all unit checkboxes based on select all state"""
        state = self.select_all_var.get()
        for var in self.unit_vars.values():
            var.set(state)

    def load_units(self, config_path):
        """Load unit configurations from config directory"""
        units = []
        for file in os.listdir(config_path):
            if file.endswith('.json'):
                with open(os.path.join(config_path, file), 'r') as f:
                    unit = json.load(f)
                    units.append(Unit(unit['unit_no'], unit['block'], 
                                    unit['ip_address'], unit['port'], 
                                    unit['serial'], unit['channels']))
        return sorted(units, key=lambda x: x.unit_no)

    def get_selected_units(self):
        return [unit for unit in self.units 
                if self.unit_vars[unit.unit_no].get()]

    def validate_date_range(self):
        try:
            start_month = list(calendar.month_name).index(self.start_month.get())
            end_month = list(calendar.month_name).index(self.end_month.get())
            start_year = int(self.start_year.get())
            end_year = int(self.end_year.get())
            
            start_date = datetime(start_year, start_month, 1)
            end_date = datetime(end_year, end_month, 
                              calendar.monthrange(end_year, end_month)[1])
            
            if start_date > end_date:
                raise ValueError("Start date must be before end date")
                
            return start_date, end_date
        except ValueError as e:
            messagebox.showerror("Error", str(e))
            return None, None

    def natural_sort_key(self, s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split('(\\d+)', s)]

    def create_temp_dir(self, prefix="maple_west_"):
        """Create a temporary directory and track it"""
        temp_dir = tempfile.mkdtemp(prefix=prefix)
        self.temp_dirs.append(temp_dir)
        return temp_dir

    def cleanup_temp_dirs(self):
        """Clean up all tracked temporary directories"""
        for dir_path in self.temp_dirs:
            if os.path.exists(dir_path):
                try:
                    shutil.rmtree(dir_path)
                except Exception as e:
                    print(f"{color.RED}Error cleaning up temporary directory {dir_path}: {str(e)}{color.END}")
        self.temp_dirs = []
        self.root.update()  # Update UI after cleanup

    def cancel_download(self):
        """Cancel the ongoing download process"""
        if self.is_downloading:
            self.is_downloading = False
            self.status_label.config(text="Download cancelled")
            self.cancel_btn.grid_remove()
            self.download_btn.config(state="normal")
            self.progress_var.set(0)  # Reset progress bar
            
            # Clean up temporary directories
            self.cleanup_temp_dirs()
            self.root.update()

    def format_time_remaining(self, seconds):
        """Format seconds into hours:minutes:seconds"""
        if seconds < 0:
            return "Calculating..."
        
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        seconds = seconds % 60
        
        if hours > 0:
            return f"{int(hours)}h {int(minutes)}m remaining"
        elif minutes > 0:
            return f"{int(minutes)}m {int(seconds)}s remaining"
        else:
            return f"{int(seconds)}s remaining"

    def update_progress(self, completed, total):
        """Update progress bar, percentage, and estimated time remaining"""
        if completed == 0:
            self.start_time = time.time()
            
        percentage = (completed / total) * 100
        self.progress_var.set(percentage)
        self.percent_label.config(text=f"{percentage:.1f}%")
        
        if completed > 0:
            elapsed_time = time.time() - self.start_time
            time_per_item = elapsed_time / completed
            remaining_items = total - completed
            estimated_time = remaining_items * time_per_item
            self.time_label.config(text=self.format_time_remaining(estimated_time))
        
        self.root.update()

    def combine_data(self, output_dir, selected_units):
        if not self.is_downloading:  # Check if cancelled
            return None
            
        self.status_label.config(text="Combining data files...")
        self.root.update()
        
        combined_dir = self.create_temp_dir(prefix="maple_west_combined_")
        
        try:
            for unit in selected_units:
                if not self.is_downloading:  # Check if cancelled
                    return None
                    
                try:
                    unit_dir = os.path.join(output_dir, f"UNIT {unit.unit_no}")
                    if not os.path.exists(unit_dir):
                        continue
                    
                    # Get all CSV files for this unit
                    csv_files = []
                    for file in os.listdir(unit_dir):
                        if file.endswith('.csv'):
                            csv_files.append(os.path.join(unit_dir, file))
                    
                    if not csv_files:
                        continue
                    
                    # Sort files by name
                    csv_files.sort(key=self.natural_sort_key)
                    
                    # Read and combine all CSV files
                    dfs = []
                    for file in csv_files:
                        try:
                            df = pd.read_csv(file)
                            if not df.empty and len(df) > 1:  # Check if we have at least 2 rows
                                try:
                                    if pd.to_datetime(df.iloc[0, 0]) > pd.to_datetime(df.iloc[1, 0]):
                                        df = df.iloc[::-1]
                                except (ValueError, TypeError) as e:
                                    print(f"{color.YELLOW}Warning: Could not parse timestamps in {file}: {str(e)}{color.END}")
                            dfs.append(df)
                        except Exception as e:
                            print(f"{color.RED}Error reading {file}: {str(e)}{color.END}")
                    
                    if dfs:
                        # Combine all DataFrames
                        combined_data = pd.concat(dfs, ignore_index=True)
                        
                        # Create unit directory in combined folder
                        combined_unit_dir = os.path.join(combined_dir, f"UNIT {unit.unit_no}")
                        os.makedirs(combined_unit_dir, exist_ok=True)
                        
                        # Save combined data
                        output_file = os.path.join(combined_unit_dir, 
                                                 f"Unit_{unit.unit_no}_combined.csv")
                        combined_data.to_csv(output_file, index=False)
                        print(f"{color.GREEN}Created combined file: {output_file}{color.END}")
                
                except Exception as e:
                    print(f"{color.RED}Error combining data for Unit {unit.unit_no}: {str(e)}{color.END}")
                    messagebox.showwarning("Warning", f"Error combining data for Unit {unit.unit_no}: {str(e)}")
            
            self.status_label.config(text="Data combination completed")
            self.root.update()
            return combined_dir
            
        except Exception as e:
            print(f"{color.RED}Error in combine_data: {str(e)}{color.END}")
            messagebox.showwarning("Warning", f"Error combining data: {str(e)}")
            return None

    def create_zip_file(self, output_dir, selected_units, data_type, start_date, end_date):
        """Create a zip file containing all downloaded data and quality reports"""
        if not os.path.exists(output_dir):
            print(f"Output directory {output_dir} does not exist")
            return None

        zip_path = os.path.join(output_dir, f'maple_west_data_{start_date}_{end_date}.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Walk through the directory and add all files
            for root, _, files in os.walk(output_dir):
                for file in files:
                    # Include CSV files (data) and Excel files (quality reports)
                    if file.endswith(('.csv', '.xlsx')):
                        file_path = os.path.join(root, file)
                        # Get the relative path for the zip file
                        rel_path = os.path.relpath(file_path, output_dir)
                        print(f"Adding {rel_path} to zip file")
                        zipf.write(file_path, rel_path)

        if os.path.exists(zip_path) and os.path.getsize(zip_path) > 0:
            return zip_path
        else:
            print(f"Error: Created zip file is empty or does not exist")
            return None

    def start_download(self):
        """
        Main download method that handles both minute and hour data downloads.
        """
        # Validate inputs
        selected_units = self.get_selected_units()
        if not selected_units:
            messagebox.showerror("Error", "Please select at least one unit")
            return
            
        start_date, end_date = self.validate_date_range()
        if not start_date or not end_date or start_date >= end_date:
            messagebox.showerror("Error", "Invalid date range")
            return
        
        # Initialize download state
        self.is_downloading = True
        self.download_btn.config(state="disabled")
        self.cancel_btn.grid()
        self.progress_var.set(0)
        self.percent_label.config(text="0%")
        self.time_label.config(text="Calculating...")
        
        try:
            data_type = self.data_type.get().capitalize()
            output_dir = self.create_temp_dir(prefix=f"maple_west_{data_type.lower()}_")
            
            # Set up date range parameters
            if data_type == "Minute":
                dates_to_download = self._get_daily_dates(start_date, end_date)
                date_format = '%Y-%m-%d'
            else:  # Hour data
                dates_to_download = self._get_monthly_dates(start_date, end_date)
                date_format = '%Y-%m'
            
            total_downloads = len(selected_units) * len(dates_to_download)
            if total_downloads == 0:
                raise ValueError("No dates to download in the specified range")
            
            # Download data
            completed = 0
            successful_downloads = 0
            downloaded_files = []  # Track successful downloads
            
            for current_date in dates_to_download:
                if not self.is_downloading:
                    break
                    
                date_str = current_date.strftime(date_format)
                
                for unit in selected_units:
                    if not self.is_downloading:
                        break
                        
                    try:
                        self.status_label.config(text=f"Downloading Unit {unit.unit_no} - {date_str}")
                        self.root.update()
                        
                        # Create unit directory
                        unit_dir = os.path.join(output_dir, f"UNIT {unit.unit_no}")
                        os.makedirs(unit_dir, exist_ok=True)
                        
                        # Download data
                        if data_type == "Minute":
                            unit.download_minute_data(date_str)
                        else:
                            unit.download_hour_data(date_str)
                            
                        # Save data if download was successful
                        if unit.data is not None and not unit.data.empty:
                            output_file = os.path.join(unit_dir, f"Unit_{unit.unit_no}_{date_str}.csv")
                            unit.data.to_csv(output_file, index=False)
                            downloaded_files.append(output_file)
                            print(f"Downloaded {output_file}")
                            successful_downloads += 1
                        else:
                            print(f"No data available for Unit {unit.unit_no} - {date_str}")
                        
                        completed += 1
                        self.update_progress(completed, total_downloads)
                        
                    except Exception as e:
                        print(f"Error downloading Unit {unit.unit_no}: {str(e)}")
            
            # Process downloaded data
            if self.is_downloading and successful_downloads > 0:
                self.status_label.config(text="Processing downloaded files...")
                
                # Combine data if selected
                if self.combine_var.get():
                    combined_dir = self.combine_data(output_dir, selected_units)
                    if combined_dir:
                        output_dir = combined_dir
                
                # Generate quality report if selected
                if self.quality_report_var.get():
                    self.generate_quality_report(output_dir, selected_units, data_type)
                
                # Create and move zip file
                self.status_label.config(text="Creating zip file...")
                temp_zip = self.create_zip_file(output_dir, selected_units, data_type,
                                              start_date.strftime('%Y%m%d'),
                                              end_date.strftime('%Y%m%d'))
                
                if temp_zip and os.path.exists(temp_zip):
                    downloads_dir = str(Path.home() / "Downloads")
                    final_zip = os.path.join(downloads_dir, os.path.basename(temp_zip))
                    shutil.move(temp_zip, final_zip)
                    
                    self.status_label.config(text="Cleaning up...")
                    self.cleanup_temp_dirs()
                    
                    messagebox.showinfo("Success",
                                      f"Process completed!\n"
                                      f"Successfully downloaded {successful_downloads} out of {total_downloads} files.\n"
                                      f"Files saved to: {final_zip}")
                    self.status_label.config(text="Ready")
                else:
                    raise ValueError("Failed to create zip file")
            else:
                if not self.is_downloading:
                    raise ValueError("Download cancelled by user")
                else:
                    raise ValueError("No data was successfully downloaded")
                    
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            print(f"Error during download: {str(e)}")
            self.status_label.config(text="Error occurred")
            
        finally:
            # Reset UI state
            self.is_downloading = False
            self.download_btn.config(state="normal")
            self.cancel_btn.grid_remove()
            self.progress_var.set(0)
            self.percent_label.config(text="0%")
            self.time_label.config(text="")
            self.cleanup_temp_dirs()
            self.root.update()

    def _get_daily_dates(self, start_date, end_date):
        """Generate list of dates for daily (minute) data downloads"""
        dates = []
        current = start_date
        while current < end_date:
            dates.append(current)
            current += timedelta(days=1)
        return dates

    def _get_monthly_dates(self, start_date, end_date):
        """Generate list of dates for monthly (hour) data downloads"""
        dates = []
        current = start_date.replace(day=1)  # Start at first of month
        end = end_date.replace(day=1)  # Compare with first of month
        while current < end:
            dates.append(current)
            current += relativedelta(months=1)
        return dates

    def generate_quality_report(self, output_dir, selected_units, data_type):
        """Generate quality report for downloaded data using QualityChecker"""
        if not self.is_downloading:  # Check if cancelled
            return
            
        self.status_label.config(text="Generating quality report...")
        self.root.update()
        
        try:
            # Initialize QualityChecker with the output directory
            checker = QualityChecker(config_path='config/', output_dir=output_dir)
            
            # Process each unit's data
            for unit in selected_units:
                unit_dir = os.path.join(output_dir, f"UNIT {unit.unit_no}")
                if not os.path.exists(unit_dir):
                    print(f"No data directory found for Unit {unit.unit_no}")
                    continue
                
                # Create DataFrame to store quality results
                report_df = pd.DataFrame(columns=[ch for ch in channels.keys() 
                                               if checker.units[unit.unit_no]['channels'].get(ch, False)])
                
                # Get all CSV files for this unit
                csv_files = []
                for file in os.listdir(unit_dir):
                    if file.endswith('.csv'):
                        csv_files.append(os.path.join(unit_dir, file))
                
                if not csv_files:
                    print(f"No CSV files found for Unit {unit.unit_no}")
                    continue
                
                # Process each file
                for file in csv_files:
                    try:
                        # Extract month from filename
                        date_str = os.path.basename(file).split('_')[-1].replace('.csv', '')
                        if len(date_str) >= 7:  # Ensure we have at least YYYY-MM
                            month = date_str[:7]  # Get YYYY-MM part
                            
                            # Process data for this month
                            results = checker.process_data(output_dir, unit.unit_no, month, data_type)
                            if results:
                                report_df.loc[month] = {ch: checker._format_quality_result(results.get(ch)) 
                                                      for ch in report_df.columns}
                    except Exception as e:
                        print(f"{color.RED}Error processing {file}: {str(e)}{color.END}")
                
                # Save results to Excel file if we have data
                if not report_df.empty:
                    report_file = os.path.join(output_dir, 'quality_reports', 
                                             f'unit_{unit.unit_no}_{data_type.lower()}_quality.xlsx')
                    os.makedirs(os.path.dirname(report_file), exist_ok=True)
                    report_df.sort_index().to_excel(report_file)
                    
                    # Apply conditional formatting
                    wb = load_workbook(report_file)
                    ws = wb.active
                    checker._apply_conditional_formatting(ws)
                    wb.save(report_file)
                    
                    print(f"{color.GREEN}Created quality report: {report_file}{color.END}")
            
            print(f"{color.GREEN}Generated quality reports in {os.path.join(output_dir, 'quality_reports')}{color.END}")
            
        except Exception as e:
            print(f"{color.RED}Error generating quality report: {str(e)}{color.END}")
            messagebox.showwarning("Warning", f"Error generating quality report: {str(e)}")

def main():
    root = tk.Tk()
    app = BulkDownloadGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main() 