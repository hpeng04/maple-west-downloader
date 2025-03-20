import os
import pandas as pd
import json
from datetime import datetime
from channels import channels
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import calendar
from rules import check_missing_rows
from unit import Unit

class QualityChecker:
    def __init__(self, config_path='config/'):
        self.units = self._load_units(config_path)
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
    def _load_units(self, config_path: str) -> list[Unit]:
        '''
        Load unit config jsons from folder path

        param: config_path: str: path to the config file
        return: list[Unit]: list of units
        '''
        units = []
        for file in os.listdir(config_path):
            if file.endswith('.json'):
                with open(os.path.join(config_path, file), 'r') as f:
                    unit = json.load(f)
                    units.append(Unit(unit['unit_no'], unit['block'], unit['ip_address'], unit['port'], unit['serial'], unit['channels']))
        return sorted(units)

    def _check_data_quality(self, data, unit, channel_name):
        

    def _format_quality_result(self, counts):
        

    def update_quality_report(self, data_type='Minute'):
        

def main():
    checker = QualityChecker()
    checker.update_quality_report('Minute')
    checker.update_quality_report('Hour')

if __name__ == "__main__":
    main()
